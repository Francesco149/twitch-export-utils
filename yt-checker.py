import os
import google.auth
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import pandas as pd
from openpyxl import load_workbook
import pickle
import sys
from config import *

# If modifying these SCOPES, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/youtube.readonly']

def authenticate_youtube():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('youtube', 'v3', credentials=creds)

def cache_name(name):
    return name + '.pickle'

def cache_load(name):
    fn = cache_name(name)
    if os.path.exists(fn):
        with open(fn, 'rb') as f:
            return pickle.load(f)
    return None

def cache_dump(name, data):
    fn = cache_name(name)
    with open(fn, 'wb') as f:
        pickle.dump(data, f)
    return data

def cached(func):
    def wrapper(*args, **kwargs):
        if c := cache_load(func.__name__):
            return c
        return cache_dump(func.__name__, func(*args, **kwargs))
    return wrapper

@cached
def fetch_all_videos(youtube):
    videos = []
    next_page_token = None

    while True:
        # Fetch a batch of videos (max 50 per request)
        request = youtube.search().list(
            part="snippet",
            forMine=True, # Include private videos
            type="video",
            maxResults=50,  # Maximum allowed per request
            pageToken=next_page_token  # Pagination token
        )
        response = request.execute()

        # Add videos to the list
        videos.extend(response['items'])

        # Check if there are more pages
        next_page_token = response.get('nextPageToken')
        if not next_page_token:
            break  # Exit the loop if there are no more pages

    return videos

def search_videos_by_timestamps(youtube, timestamps):
    # Fetch all videos from the channel (with pagination)
    videos = [x for x in fetch_all_videos(youtube) if 'snippet' in x]
    print(f"found {len(videos)} videos")

    # Match timestamps to video titles
    video_urls = []
    for timestamp in timestamps:
        found = False
        for video in videos:
            if timestamp in video['snippet']['title']:
                video_id = video['id']['videoId']
                video_urls.append(f"https://www.youtube.com/watch?v={video_id}")
                found = True
                break
        if not found:
            video_urls.append("")  # No match found
    return video_urls

def read_timestamps_from_excel(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    timestamps = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0]:
            timestamps.append(row[0])
    return timestamps

def main():
    # Path to the Excel file
    excel_file_path = SPREADSHEET

    # Read timestamps from the Excel file
    try:
        timestamps = read_timestamps_from_excel(excel_file_path)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Authenticate and initialize the YouTube API client
    youtube = authenticate_youtube()

    # Search for videos and collect URLs
    video_urls = search_videos_by_timestamps(youtube, timestamps)

    # used to temporarily mark long vods in the spreadsheet so I don't
    # click them
    if sys.argv[1] == "-m":
        with open("long_vods.pkl", "rb") as file:
            longvods = pickle.load(file)

        longvods_ts = [title.split(' ')[0] for (url, title) in longvods]

        for i, (ts, url) in enumerate(zip(timestamps, video_urls)):
          if ts in longvods_ts:
              video_urls[i] = "*** LONG VOD >12h ***"

    # Create a DataFrame with the results
    df = pd.DataFrame({
        'Timestamp': timestamps,
        'Video URL': video_urls
    })

    # Save the results to a CSV file
    output_csv_path = 'output.csv'
    df.to_csv(output_csv_path, index=False)
    print(f"Results saved to {output_csv_path}")

if __name__ == '__main__':
    main()
